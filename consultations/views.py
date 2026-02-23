'\nОбращения (requests), консультации, отчёты. Доступ: психолог, администратор.\n'
import json
import os
import re
import shutil
import subprocess
from datetime import timedelta
from pathlib import Path
from django.db.models import Q, Count, Max, Sum, Avg, OuterRef, Subquery, IntegerField, Value
from django.db.models.functions import TruncMonth, Coalesce
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View,
)
from django.contrib import messages
from django.utils import timezone

from users.decorators import PsychologistRequiredMixin, AdminRequiredMixin, StudentRequiredMixin
from students.models import Student
from .models import (
    Request,
    Consultation,
    RequestStatus,
    ConsultationStudent,
    StudentNotification,
    Attachment,
    Log,
    Note,
    RequestNote,
    StudentPsychologistChat,
    ChatMessage,
    ChatMessageRead,
)
from .forms import (
    RequestForm,
    ConsultationForm,
    ConsultationNoteForm,
    MyRequestCreateForm,
    ChatMessageForm,
    ConsultationPsychologistAssignForm,
)
from .signals import notify_request_status_changed


def _resolve_pg_dump_path():
    """Возвращает путь к pg_dump: settings/env -> PATH -> стандартные пути Windows."""
    configured = getattr(settings, 'PG_DUMP_PATH', None) or os.environ.get('PG_DUMP_PATH')
    if configured:
        configured_path = Path(configured)
        if configured_path.exists():
            return str(configured_path)
    from_path = shutil.which('pg_dump')
    if from_path:
        return from_path

    # Windows fallback: C:\Program Files\PostgreSQL\<version>\bin\pg_dump.exe
    pf = os.environ.get('ProgramFiles', r'C:\Program Files')
    root = Path(pf) / 'PostgreSQL'
    if root.exists():
        candidates = []
        for ver_dir in root.iterdir():
            exe = ver_dir / 'bin' / 'pg_dump.exe'
            if exe.exists():
                candidates.append(exe)
        if candidates:
            candidates.sort(reverse=True)
            return str(candidates[0])
    return None


def _get_pdf_cyrillic_font():
    'Регистрирует и возвращает имя шрифта с поддержкой кириллицы для reportlab.'
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = 'CyrillicPDFFont'
    try:
        if pdfmetrics.getFont(font_name) is not None:
            return font_name
    except Exception:
        pass
    # Windows: Arial
    win_path = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'arial.ttf')
    if os.path.isfile(win_path):
        pdfmetrics.registerFont(TTFont(font_name, win_path))
        return font_name
    # Linux: DejaVu Sans
    for path in ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                 '/usr/share/fonts/TTF/DejaVuSans.ttf'):
        if os.path.isfile(path):
            pdfmetrics.registerFont(TTFont(font_name, path))
            return font_name
    # Fallback: Helvetica (кириллица не отобразится, но PDF не упадёт)
    return 'Helvetica'


# ——— Обращения (requests) ———

class RequestListView(PsychologistRequiredMixin, ListView):
    model = Request
    template_name = 'consultations/request_list.html'
    context_object_name = 'requests'
    paginate_by = 0  # не используем общую пагинацию — разбивка по статусам

    def get_queryset(self):
        qs = Request.objects.select_related('student', 'status').order_by('-created_at')
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(student__last_name__icontains=q)
                | Q(student__first_name__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search_q'] = self.request.GET.get('q', '')
        qs = self.get_queryset()
        ctx['requests_new'] = qs.filter(status__name='new')[:50]
        ctx['requests_in_progress'] = qs.filter(status__name='in_progress')[:50]
        ctx['requests_completed'] = qs.filter(status__name='completed')[:50]
        ctx['requests_cancelled'] = qs.filter(status__name='cancelled')[:50]
        return ctx


class RequestDetailView(PsychologistRequiredMixin, DetailView):
    model = Request
    template_name = 'consultations/request_detail.html'
    context_object_name = 'request_obj'

    def get_queryset(self):
        return Request.objects.select_related('student', 'status').prefetch_related('consultations__form')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        req = ctx.get('request_obj')
        ctx['request_notes'] = (
            RequestNote.objects.filter(request_id=req.pk).select_related('user').order_by('created_at')
            if req else []
        )
        return ctx


class RequestCreateView(PsychologistRequiredMixin, CreateView):
    model = Request
    form_class = RequestForm
    template_name = 'consultations/request_form.html'
    success_url = reverse_lazy('consultations:request_list')

    def get_initial(self):
        initial = super().get_initial()
        sid = self.request.GET.get('student')
        if sid:
            try:
                initial['student'] = Student.objects.get(pk=sid)
            except (ValueError, Student.DoesNotExist):
                pass
        new_status = RequestStatus.objects.filter(name='new').first()
        if new_status:
            initial['status'] = new_status
        return initial

    def form_valid(self, form):
        form.instance.psychologist = self.request.user
        messages.success(self.request, 'Обращение зарегистрировано.')
        return super().form_valid(form)


class RequestUpdateView(PsychologistRequiredMixin, UpdateView):
    model = Request
    form_class = RequestForm
    template_name = 'consultations/request_form.html'
    success_url = reverse_lazy('consultations:request_list')
    context_object_name = 'request_obj'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Администратор не может изменять обращения.')
            return redirect('consultations:request_detail', pk=kwargs.get('pk'))
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Обращение обновлено.')
        return super().form_valid(form)


class RequestDeleteView(AdminRequiredMixin, DeleteView):
    model = Request
    template_name = 'consultations/request_confirm_delete.html'
    success_url = reverse_lazy('consultations:request_list')
    context_object_name = 'request_obj'

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Обращение удалено.')
        return super().delete(request, *args, **kwargs)


class RequestCompleteView(PsychologistRequiredMixin, View):
    'Завершение обращения (только психолог): статус переводится в «Завершено».'
    def post(self, request, pk):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Только психолог может завершать обращения.')
            return redirect('consultations:request_detail', pk=pk)
        request_obj = get_object_or_404(Request, pk=pk)
        completed_status = RequestStatus.objects.filter(name='completed').first()
        if not completed_status:
            messages.error(request, 'В системе не найден статус «Завершено».')
            return redirect('consultations:request_detail', pk=pk)
        if request_obj.status.name == 'completed':
            messages.info(request, 'Обращение уже завершено.')
            return redirect('consultations:request_detail', pk=pk)
        request_obj.status = completed_status
        request_obj.save(update_fields=['status_id'])
        notify_request_status_changed(request_obj)
        messages.success(request, 'Обращение завершено.')
        return redirect('consultations:request_detail', pk=pk)


class RequestCancelView(PsychologistRequiredMixin, View):
    'Отмена обращения (только психолог, до статуса «Завершено»).'
    def post(self, request, pk):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Только психолог может отменять обращения.')
            return redirect('consultations:request_detail', pk=pk)
        request_obj = get_object_or_404(Request, pk=pk)
        if request_obj.status.name == 'completed':
            messages.warning(request, 'Завершённое обращение отменить нельзя.')
            return redirect('consultations:request_detail', pk=pk)
        cancelled_status = RequestStatus.objects.filter(name='cancelled').first()
        if not cancelled_status:
            messages.error(request, 'В системе не найден статус «Отменено». Обратитесь к администратору.')
            return redirect('consultations:request_detail', pk=pk)
        request_obj.status = cancelled_status
        request_obj.save(update_fields=['status_id'])
        notify_request_status_changed(request_obj)
        messages.success(request, 'Обращение отменено.')
        return redirect('consultations:request_list')


# ——— Консультации ———

class ConsultationListView(PsychologistRequiredMixin, ListView):
    model = Consultation
    template_name = 'consultations/consultation_list.html'
    context_object_name = 'consultations'
    paginate_by = 20

    def get_queryset(self):
        qs = Consultation.objects.select_related('request', 'request__student', 'form').prefetch_related('students').order_by('-date', '-created_at')
        date_from = self.request.GET.get('date_from', '').strip()
        date_to = self.request.GET.get('date_to', '').strip()
        student_id = self.request.GET.get('student', '').strip()
        q = self.request.GET.get('q', '').strip()
        tab = self.request.GET.get('tab', '').strip()  # '' | 'upcoming' | 'past'

        today = timezone.now().date()
        if tab == 'upcoming':
            # Запланированные: дата в будущем или сегодня и ещё не завершена
            qs = qs.filter(date__gte=today, completed_at__isnull=True)
        elif tab == 'past':
            # Прошедшие: дата в прошлом или консультация завершена (результаты в отчётности)
            qs = qs.filter(Q(date__lt=today) | Q(completed_at__isnull=False))

        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if student_id:
            qs = qs.filter(Q(request__student_id=student_id) | Q(students__id=student_id)).distinct()
        if q:
            qs = qs.filter(
                Q(request__student__last_name__icontains=q)
                | Q(request__student__first_name__icontains=q)
                | Q(students__last_name__icontains=q)
                | Q(students__first_name__icontains=q)
                | Q(result__icontains=q)
            ).distinct()
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search_q'] = self.request.GET.get('q', '')
        ctx['date_from'] = self.request.GET.get('date_from', '')
        ctx['date_to'] = self.request.GET.get('date_to', '')
        ctx['today'] = timezone.now().date()
        ctx['tab'] = self.request.GET.get('tab', '')
        sid = self.request.GET.get('student', '')
        ctx['filter_student'] = sid
        try:
            ctx['filter_student_id'] = int(sid) if sid else None
        except ValueError:
            ctx['filter_student_id'] = None
        ctx['students'] = Student.objects.all().order_by('last_name', 'first_name')[:200]
        return ctx


class ConsultationDetailView(PsychologistRequiredMixin, DetailView):
    model = Consultation
    template_name = 'consultations/consultation_detail.html'
    context_object_name = 'consultation'

    def get_queryset(self):
        return Consultation.objects.select_related('request', 'form').prefetch_related(
            'students', 'consultation_students__student', 'attachments', 'notes__user'
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['today'] = timezone.now().date()
        consultation = ctx.get('consultation')
        links = list(consultation.consultation_students.all()) if consultation else []
        active_links = [link for link in links if not link.participation_cancelled_at]
        ctx['all_participation_confirmed'] = (
            all(link.participation_confirmed_at for link in active_links)
            if active_links else True
        )
        ctx['assigned_psychologist'] = (
            consultation.request.psychologist if consultation and consultation.request_id else None
        )
        if self.request.user.role_name == 'admin':
            initial = {}
            if consultation and consultation.request_id and consultation.request.psychologist_id:
                initial['psychologist'] = consultation.request.psychologist_id
            ctx['assign_psychologist_form'] = ConsultationPsychologistAssignForm(initial=initial)
        ctx['consultation_notes'] = (
            consultation.notes.select_related('user').order_by('created_at')
            if consultation else []
        )
        if self.request.user.role_name == 'psychologist':
            ctx['consultation_note_form'] = ConsultationNoteForm()
        return ctx


class ConsultationNoteCreateView(PsychologistRequiredMixin, View):
    """Добавление заметки психологом в карточке консультации."""
    def post(self, request, pk):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Только психолог может добавлять заметки к консультации.')
            return redirect('consultations:consultation_detail', pk=pk)

        consultation = get_object_or_404(Consultation, pk=pk)
        form = ConsultationNoteForm(request.POST)
        if not form.is_valid():
            error_text = '; '.join(
                [err for errs in form.errors.values() for err in errs]
            ) or 'Проверьте корректность заметки.'
            messages.error(request, error_text)
            return redirect('consultations:consultation_detail', pk=pk)

        Note.objects.create(
            consultation_id=consultation.pk,
            user_id=request.user.pk,
            text=form.cleaned_data['text'],
        )
        messages.success(request, 'Заметка к консультации добавлена.')
        return redirect('consultations:consultation_detail', pk=pk)


class ConsultationAssignPsychologistView(AdminRequiredMixin, View):
    'Назначение психолога к обращению, связанному с консультацией (только админ).'
    def post(self, request, pk):
        consultation = get_object_or_404(Consultation.objects.select_related('request'), pk=pk)
        if not consultation.request_id:
            messages.error(request, 'Для этой консультации не найдено связанное обращение.')
            return redirect('consultations:consultation_detail', pk=pk)
        form = ConsultationPsychologistAssignForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Выберите психолога из списка.')
            return redirect('consultations:consultation_detail', pk=pk)
        assigned = form.cleaned_data['psychologist']
        req = consultation.request
        if req.psychologist_id == assigned.pk:
            messages.info(request, f'Психолог уже установлен: {assigned.username}.')
            return redirect('consultations:consultation_detail', pk=pk)
        req.psychologist_id = assigned.pk
        req.save(update_fields=['psychologist_id'])
        messages.success(request, f'Психолог для консультации установлен: {assigned.username}.')
        return redirect('consultations:consultation_detail', pk=pk)


class ConsultationCreateView(PsychologistRequiredMixin, CreateView):
    model = Consultation
    form_class = ConsultationForm
    template_name = 'consultations/consultation_form.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Администратор не может регистрировать консультации.')
            return redirect('consultations:consultation_list')
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        req_id = self.request.GET.get('request')
        if req_id:
            req = get_object_or_404(Request, pk=req_id)
            initial['request'] = req
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        req_id = self.request.GET.get('request')
        linked_request = None
        if req_id:
            try:
                linked_request = Request.objects.select_related('student').get(pk=req_id)
            except Request.DoesNotExist:
                linked_request = None
        ctx['linked_request'] = linked_request
        return ctx

    def get_success_url(self):
        # После регистрации консультации по обращению — обратно к карточке обращения
        req_id = getattr(self, '_created_request_id', None)
        if req_id:
            return reverse_lazy('consultations:request_detail', kwargs={'pk': req_id})
        return reverse_lazy('consultations:consultation_list')

    def form_valid(self, form):
        consultation = form.save()
        self._created_request_id = consultation.request_id
        req = consultation.request
        # Психолог, зарегистрировавший консультацию, становится ведущим по обращению.
        if req and req.psychologist_id != self.request.user.id:
            req.psychologist_id = self.request.user.id
            req.save(update_fields=['psychologist_id'])
        # При первой консультации по обращению переводим статус в «В работе»
        if req and req.status.name == 'new':
            in_progress = RequestStatus.objects.filter(name='in_progress').first()
            if in_progress:
                req.status = in_progress
                req.save(update_fields=['status_id'])
                notify_request_status_changed(req)
        messages.success(self.request, 'Консультация зарегистрирована.')
        return redirect(self.get_success_url())


class ConsultationUpdateView(PsychologistRequiredMixin, UpdateView):
    model = Consultation
    form_class = ConsultationForm
    template_name = 'consultations/consultation_form.html'
    context_object_name = 'consultation'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Администратор не может изменять сведения консультации.')
            return redirect('consultations:consultation_detail', pk=kwargs.get('pk'))
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('consultations:consultation_list')

    def form_valid(self, form):
        messages.success(self.request, 'Консультация обновлена.')
        return super().form_valid(form)


class ConsultationDeleteView(PsychologistRequiredMixin, DeleteView):
    model = Consultation
    template_name = 'consultations/consultation_confirm_delete.html'
    success_url = reverse_lazy('consultations:consultation_list')
    context_object_name = 'consultation'

    def dispatch(self, request, *args, **kwargs):
        consultation = get_object_or_404(Consultation, pk=kwargs.get('pk'))
        if not consultation.completed_at and not consultation.cancelled_at:
            messages.error(request, 'Удаление возможно только для завершённой или отменённой консультации.')
            return redirect('consultations:consultation_detail', pk=consultation.pk)
        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Консультация удалена.')
        return super().delete(request, *args, **kwargs)


class DatabaseMaintenanceView(AdminRequiredMixin, TemplateView):
    """Резервное копирование PostgreSQL для администратора."""
    template_name = 'consultations/database_maintenance.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role_name != 'admin':
            raise PermissionDenied('Доступ разрешён только администратору.')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        backup_dir = Path(settings.BACKUP_DIR)
        backup_dir.mkdir(parents=True, exist_ok=True)
        backups = []
        files = list(backup_dir.glob('backup_*.sql')) + list(backup_dir.glob('backup_*.dump'))
        for f in sorted(files, key=lambda p: p.stat().st_mtime, reverse=True):
            stat = f.stat()
            backups.append({
                'name': f.name,
                'format': f.suffix.lower().lstrip('.'),
                'size_mb': round(stat.st_size / (1024 * 1024), 2),
                'modified_at': timezone.datetime.fromtimestamp(stat.st_mtime, tz=timezone.get_current_timezone()),
            })
        ctx['backups'] = backups
        return ctx

    def post(self, request, *args, **kwargs):
        db = settings.DATABASES.get('default', {})
        engine = (db.get('ENGINE') or '').lower()
        if 'postgresql' not in engine:
            messages.error(request, 'Резервное копирование доступно только для PostgreSQL.')
            return redirect('consultations:database_maintenance')

        pg_dump = _resolve_pg_dump_path()
        if not pg_dump:
            messages.error(
                request,
                'pg_dump недоступен. Укажите PG_DUMP_PATH в settings.py или добавьте PostgreSQL bin в PATH.'
            )
            return redirect('consultations:database_maintenance')

        backup_dir = Path(settings.BACKUP_DIR)
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_format = (request.POST.get('backup_format') or 'sql').strip().lower()
        if backup_format not in {'sql', 'dump'}:
            backup_format = 'sql'
        ext = 'sql' if backup_format == 'sql' else 'dump'
        filename = f"backup_{timezone.now().strftime('%Y-%m-%d_%H-%M-%S')}.{ext}"
        backup_path = backup_dir / filename

        cmd = [
            pg_dump,
            '-h', str(db.get('HOST') or '127.0.0.1'),
            '-p', str(db.get('PORT') or '5432'),
            '-U', str(db.get('USER') or ''),
            '-d', str(db.get('NAME') or ''),
            '-F', 'p' if backup_format == 'sql' else 'c',
            '--encoding=UTF8',
            '--clean',
            '--if-exists',
            '--no-owner',
            '--no-privileges',
            '-f', str(backup_path),
        ]
        env = os.environ.copy()
        env['PGPASSWORD'] = str(db.get('PASSWORD') or '')
        proc = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='ignore', env=env)

        if proc.returncode != 0:
            messages.error(request, f'Не удалось создать бэкап: {(proc.stderr or "").strip()[:300]}')
            return redirect('consultations:database_maintenance')

        Log.objects.create(user_id=request.user.pk, action='Backup created')
        messages.success(request, f'Бэкап успешно создан: {filename}')
        return redirect('consultations:database_maintenance')


class DatabaseBackupDownloadView(AdminRequiredMixin, View):
    """Скачивание созданного backup-файла."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.role_name != 'admin':
            raise PermissionDenied('Доступ разрешён только администратору.')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, filename):
        if not re.fullmatch(r'backup_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}\.(sql|dump)', filename):
            raise Http404('Файл не найден.')
        backup_dir = Path(settings.BACKUP_DIR).resolve()
        file_path = (backup_dir / filename).resolve()
        if backup_dir not in file_path.parents and file_path != backup_dir:
            raise Http404('Файл не найден.')
        if not file_path.exists() or not file_path.is_file():
            raise Http404('Файл не найден.')
        response = FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=filename,
            content_type='application/octet-stream',
        )
        response['X-Content-Type-Options'] = 'nosniff'
        return response


class DatabaseBackupDeleteView(AdminRequiredMixin, View):
    """Удаление созданного backup-файла."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.role_name != 'admin':
            raise PermissionDenied('Доступ разрешён только администратору.')
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, filename):
        if not re.fullmatch(r'backup_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}\.(sql|dump)', filename):
            raise Http404('Файл не найден.')
        backup_dir = Path(settings.BACKUP_DIR).resolve()
        file_path = (backup_dir / filename).resolve()
        if backup_dir not in file_path.parents and file_path != backup_dir:
            raise Http404('Файл не найден.')
        if not file_path.exists() or not file_path.is_file():
            messages.error(request, 'Файл бэкапа не найден.')
            return redirect('consultations:database_maintenance')
        try:
            file_path.unlink()
        except OSError:
            messages.error(request, 'Не удалось удалить файл бэкапа.')
            return redirect('consultations:database_maintenance')
        Log.objects.create(user_id=request.user.pk, action='Backup deleted')
        messages.success(request, f'Бэкап удалён: {filename}')
        return redirect('consultations:database_maintenance')


def _safe_attachment_filename(name):
    'Оставляет только безопасные символы в имени файла.'
    import re
    import uuid
    base, *rest = name.rsplit('.', 1)
    ext = ('.' + rest[0]) if rest else ''
    safe_base = re.sub(r'[^\w\-]', '_', base)[:80] or uuid.uuid4().hex[:8]
    return f'{safe_base}_{uuid.uuid4().hex[:8]}{ext}'


class ConsultationAttachmentUploadView(PsychologistRequiredMixin, View):
    'Прикрепление документа к результату консультации.'
    def post(self, request, pk):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Администратор не может прикреплять документы.')
            return redirect('consultations:consultation_detail', pk=pk)
        from django.conf import settings
        import os
        consultation = get_object_or_404(Consultation, pk=pk)
        f = request.FILES.get('file')
        if not f:
            messages.error(request, 'Выберите файл для загрузки.')
            return redirect('consultations:consultation_detail', pk=pk)
        rel_dir = os.path.join('consultations', str(consultation.pk))
        dest_dir = os.path.join(settings.MEDIA_ROOT, rel_dir)
        os.makedirs(dest_dir, exist_ok=True)
        filename = _safe_attachment_filename(f.name)
        rel_path = os.path.join(rel_dir, filename)
        full_path = os.path.join(settings.MEDIA_ROOT, rel_path)
        try:
            with open(full_path, 'wb') as out:
                for chunk in f.chunks():
                    out.write(chunk)
        except Exception as e:
            messages.error(request, f'Ошибка сохранения файла: {e}')
            return redirect('consultations:consultation_detail', pk=pk)
        description = (request.POST.get('description') or '').strip() or None
        Attachment.objects.create(
            consultation_id=consultation.pk,
            file_path=rel_path.replace('\\', '/'),
            description=description,
        )
        messages.success(request, 'Документ прикреплён.')
        return redirect('consultations:consultation_detail', pk=pk)


class ConsultationAttachmentDeleteView(PsychologistRequiredMixin, View):
    'Удаление вложения консультации.'
    def post(self, request, pk, attachment_id):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Администратор не может удалять документы.')
            return redirect('consultations:consultation_detail', pk=pk)
        from django.conf import settings
        import os
        consultation = get_object_or_404(Consultation, pk=pk)
        att = get_object_or_404(Attachment, pk=attachment_id, consultation_id=consultation.pk)
        if att.file_path:
            full_path = os.path.join(settings.MEDIA_ROOT, att.file_path.replace('/', os.sep))
            if os.path.isfile(full_path):
                try:
                    os.remove(full_path)
                except OSError:
                    pass
        att.delete()
        messages.success(request, 'Вложение удалено.')
        return redirect('consultations:consultation_detail', pk=pk)


class ConsultationCompleteView(PsychologistRequiredMixin, View):
    'Завершение консультации: помечаем как пройденную, результаты попадают в отчётность.'
    def post(self, request, pk):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Только психолог может завершать консультации.')
            return redirect('consultations:consultation_detail', pk=pk)
        consultation = get_object_or_404(
            Consultation.objects.select_related('request').prefetch_related('consultation_students'),
            pk=pk
        )
        if not consultation.request_id or not consultation.request.psychologist_id:
            messages.warning(request, 'Нельзя завершить консультацию, пока не назначен психолог.')
            return redirect('consultations:consultation_detail', pk=pk)
        if consultation.request.psychologist_id != request.user.id:
            messages.warning(request, 'Завершить консультацию может только закреплённый за ней психолог.')
            return redirect('consultations:consultation_detail', pk=pk)
        if consultation.completed_at:
            messages.info(request, 'Консультация уже завершена.')
            return redirect('consultations:consultation_detail', pk=pk)
        if consultation.cancelled_at:
            messages.warning(request, 'Консультация отменена. Завершение невозможно.')
            return redirect('consultations:consultation_detail', pk=pk)
        if not (consultation.result or '').strip():
            messages.warning(
                request,
                'Для завершения консультации укажите результат. Нажмите «Изменить» и заполните поле «Результат».'
            )
            return redirect('consultations:consultation_edit', pk=pk)
        # Требуем подтверждение участия всеми учащимися (кто не отменил участие окончательно)
        links = list(consultation.consultation_students.all())
        if links:
            active = [link for link in links if not link.participation_cancelled_at]
            unconfirmed = [link for link in active if not link.participation_confirmed_at]
            if unconfirmed:
                names = ', '.join(link.student.full_name for link in unconfirmed)
                messages.warning(
                    request,
                    f'Завершение невозможно: участие не подтверждено учащимися: {names}. '
                    'Учащийся должен подтвердить участие в разделе «Мои консультации».'
                )
                return redirect('consultations:consultation_detail', pk=pk)
        consultation.completed_at = timezone.now()
        consultation.save(update_fields=['completed_at'])
        # Обращение считаем завершённым после завершения консультации по нему
        if consultation.request_id:
            completed_status = RequestStatus.objects.filter(name='completed').first()
            if completed_status:
                consultation.request.status = completed_status
                consultation.request.save(update_fields=['status_id'])
                notify_request_status_changed(consultation.request)
        messages.success(request, 'Консультация завершена. Результаты учтены в отчётности.')
        return redirect('consultations:consultation_detail', pk=pk)


class ConsultationCancelView(PsychologistRequiredMixin, View):
    'Отмена консультации (только психолог).'
    def post(self, request, pk):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Только психолог может отменять консультации.')
            return redirect('consultations:consultation_detail', pk=pk)
        consultation = get_object_or_404(Consultation.objects.select_related('request'), pk=pk)
        if consultation.completed_at:
            messages.warning(request, 'Завершённую консультацию отменить нельзя.')
            return redirect('consultations:consultation_detail', pk=pk)
        if consultation.cancelled_at:
            messages.info(request, 'Консультация уже отменена.')
            return redirect('consultations:consultation_detail', pk=pk)
        consultation.cancelled_at = timezone.now()
        consultation.save(update_fields=['cancelled_at'])
        # Обращение по этой консультации тоже переводим в «Отменённые»
        if consultation.request_id:
            cancelled_status = RequestStatus.objects.filter(name='cancelled').first()
            if cancelled_status:
                consultation.request.status = cancelled_status
                consultation.request.save(update_fields=['status_id'])
                notify_request_status_changed(consultation.request)
        messages.success(request, 'Консультация отменена. Связанное обращение переведено в статус «Отменено».')
        return redirect('consultations:consultation_detail', pk=pk)


# ——— Отчёты ———

def _report_filters(request):
    'Общие фильтры для отчётов (даты, статус, учащийся).'
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    status = request.GET.get('status', '').strip()
    student_id = request.GET.get('student', '').strip()
    return date_from, date_to, status, student_id


def _apply_report_filters(qs_req, qs_cons, date_from, date_to, status, student_id):
    'Применяет фильтры к QuerySet обращений и консультаций.'
    if date_from:
        qs_req = qs_req.filter(created_at__date__gte=date_from)
        # РљРѕРЅСЃСѓР»СЊС‚Р°С†РёСЏ РІ РїРµСЂРёРѕРґРµ: РїРѕ РґР°С‚Рµ РїСЂРѕРІРµРґРµРЅРёСЏ РР›Р РїРѕ РґР°С‚Рµ Р·Р°РІРµСЂС€РµРЅРёСЏ
        qs_cons = qs_cons.filter(Q(date__gte=date_from) | Q(completed_at__date__gte=date_from))
    if date_to:
        qs_req = qs_req.filter(created_at__date__lte=date_to)
        qs_cons = qs_cons.filter(Q(date__lte=date_to) | Q(completed_at__date__lte=date_to))
    if status:
        qs_req = qs_req.filter(status__name=status)
    if student_id:
        qs_req = qs_req.filter(student_id=student_id)
        qs_cons = qs_cons.filter(Q(request__student_id=student_id) | Q(students__id=student_id)).distinct()
    return qs_req, qs_cons


class ReportView(PsychologistRequiredMixin, TemplateView):
    template_name = 'consultations/report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        is_admin = user.role_name == 'admin'
        is_psychologist = user.role_name == 'psychologist'

        date_from, date_to, status, student_id = _report_filters(self.request)

        # Базовые queryset: у психолога — свои обращения и консультации (в т.ч. без обращения)
        qs_req = Request.objects.all()
        qs_cons = Consultation.objects.select_related('request')
        if is_psychologist:
            # Обращения: созданные учащимися (без психолога) + назначенные этому психологу
            qs_req = qs_req.filter(Q(psychologist_id=user.id) | Q(psychologist_id__isnull=True))
            # Свои консультации: по своим обращениям, без обращения или по обращениям без привязки психолога
            qs_cons = qs_cons.filter(
                Q(request__psychologist_id=user.id)
                | Q(request_id__isnull=True)
                | Q(request__psychologist_id__isnull=True)
            )

        qs_req, qs_cons = _apply_report_filters(qs_req, qs_cons, date_from, date_to, status, student_id)

        ctx['is_admin'] = is_admin
        ctx['is_psychologist'] = is_psychologist
        ctx['date_from'] = date_from
        ctx['date_to'] = date_to
        ctx['filter_status'] = status
        ctx['filter_student'] = student_id
        try:
            ctx['filter_student_id'] = int(student_id) if student_id else None
        except ValueError:
            ctx['filter_student_id'] = None
        ctx['students'] = Student.objects.all().order_by('last_name', 'first_name')[:300]

        # ——— Для психолога: три отчёта ———
        if is_psychologist:
            student_ids = set(qs_req.values_list('student_id', flat=True))
            student_ids |= set(qs_cons.values_list('request__student_id', flat=True))
            student_ids |= set(qs_cons.values_list('students__id', flat=True))
            student_ids.discard(None)
            students_list = list(Student.objects.filter(pk__in=student_ids).order_by('last_name', 'first_name')) if student_ids else []

            # 1) Обращения и консультации по учащимся
            students_report_data = []
            for s in students_list:
                req_cnt = qs_req.filter(student_id=s.id).count()
                cons_qs = qs_cons.filter(Q(request__student_id=s.id) | Q(students__id=s.id)).distinct()
                cons_cnt = cons_qs.count()
                last_cons = cons_qs.aggregate(m=Max('date'))['m']
                students_report_data.append({
                    'student': s,
                    'request_count': req_cnt,
                    'consultation_count': cons_cnt,
                    'last_consultation_date': last_cons,
                })
            ctx['students_report_data'] = students_report_data

            # 2) Динамика обращений по месяцам (новые, в работе, завершённые, отменённые)
            request_dynamics_list = _get_request_dynamics_data(qs_req)
            ctx['request_dynamics'] = request_dynamics_list
            ctx['request_dynamics_chart_json'] = json.dumps([{
                'label': d['label'],
                'req_new': d['req_new'],
                'req_in_progress': d['req_in_progress'],
                'req_completed': d['req_completed'],
                'req_cancelled': d['req_cancelled'],
                'cnt': d['cnt']
            } for d in request_dynamics_list]) if request_dynamics_list else '[]'
            # 3) Динамика консультаций по месяцам (завершённые, отменённые)
            consultation_dynamics_list = _get_consultation_dynamics_data(qs_cons)
            ctx['consultation_dynamics'] = consultation_dynamics_list
            ctx['consultation_dynamics_chart_json'] = json.dumps([{
                'label': d['label'],
                'cons_completed': d['cons_completed'],
                'cons_cancelled': d['cons_cancelled'],
                'cnt': d['cnt']
            } for d in consultation_dynamics_list]) if consultation_dynamics_list else '[]'

            # 4) Нагрузка школьного психолога
            ctx['workload_requests'] = qs_req.count()
            ctx['workload_consultations'] = qs_cons.count()
            completed_cons = qs_cons.filter(completed_at__isnull=False)
            dur_agg = completed_cons.aggregate(total=Sum('duration'), avg=Avg('duration'))
            ctx['workload_duration_total'] = dur_agg['total'] or 0
            ctx['workload_duration_avg'] = round(dur_agg['avg'], 1) if dur_agg['avg'] is not None else 0

        # ——— Для админа: статистика по психологам и общая, отчёт по консультациям ———
        if is_admin:
            ctx['consultation_count'] = qs_cons.count()
            # По психологам: количество обращений (те же фильтры: даты, статус, учащийся)
            by_psychologist = Request.objects.filter(psychologist_id__isnull=False)
            if date_from:
                by_psychologist = by_psychologist.filter(created_at__date__gte=date_from)
            if date_to:
                by_psychologist = by_psychologist.filter(created_at__date__lte=date_to)
            if status:
                by_psychologist = by_psychologist.filter(status__name=status)
            if student_id:
                by_psychologist = by_psychologist.filter(student_id=student_id)
            by_psychologist = list(
                by_psychologist.values('psychologist_id', 'psychologist__username')
                .annotate(cnt=Count('id')).order_by('-cnt')
            )
            for r in by_psychologist:
                r['name'] = r.get('psychologist__username', '—')
            ctx['request_by_psychologist'] = by_psychologist
            # Общая статистика обращений (те же фильтры)
            req_all = Request.objects.all()
            if date_from:
                req_all = req_all.filter(created_at__date__gte=date_from)
            if date_to:
                req_all = req_all.filter(created_at__date__lte=date_to)
            if status:
                req_all = req_all.filter(status__name=status)
            if student_id:
                req_all = req_all.filter(student_id=student_id)
            ctx['request_total'] = req_all.count()
            _status_labels = {'new': 'Новое', 'in_progress': 'В работе', 'completed': 'Завершено', 'cancelled': 'Отменено'}
            req_by_status = list(req_all.values('status__name').annotate(cnt=Count('id')).order_by('-cnt'))
            for r in req_by_status:
                r['status_display'] = _status_labels.get(r.get('status__name'), r.get('status__name') or '—')
            ctx['request_by_status'] = req_by_status
            # Список консультаций для отчёта (с фильтрами)
            ctx['consultations_in_report'] = qs_cons.select_related('request__student', 'request__psychologist').prefetch_related('students').order_by('-date')[:500]

        return ctx


class StudentDynamicsView(PsychologistRequiredMixin, TemplateView):
    template_name = 'consultations/student_dynamics.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        student = get_object_or_404(
            Student.objects.select_related('classroom'),
            pk=self.kwargs.get('pk'),
        )

        date_from, date_to, _, _ = _report_filters(self.request)

        req_qs = Request.objects.filter(student_id=student.id).select_related('status', 'psychologist')
        cons_qs = Consultation.objects.filter(
            Q(request__student_id=student.id) | Q(students__id=student.id)
        ).select_related('request', 'request__status', 'form').prefetch_related('students').distinct()

        if user.role_name == 'psychologist':
            req_qs = req_qs.filter(Q(psychologist_id=user.id) | Q(psychologist_id__isnull=True))
            cons_qs = cons_qs.filter(
                Q(request__psychologist_id=user.id)
                | Q(request_id__isnull=True)
                | Q(request__psychologist_id__isnull=True)
            )

        if date_from:
            req_qs = req_qs.filter(created_at__date__gte=date_from)
            cons_qs = cons_qs.filter(date__gte=date_from)
        if date_to:
            req_qs = req_qs.filter(created_at__date__lte=date_to)
            cons_qs = cons_qs.filter(date__lte=date_to)

        if user.role_name == 'psychologist' and not req_qs.exists() and not cons_qs.exists():
            raise PermissionDenied('Нет доступа к аналитике этого учащегося.')

        req_total = req_qs.count()
        req_status_map = {'new': 0, 'in_progress': 0, 'completed': 0, 'cancelled': 0}
        for row in req_qs.values('status__name').annotate(cnt=Count('id')):
            name = row.get('status__name')
            if name in req_status_map:
                req_status_map[name] = row.get('cnt', 0) or 0

        cons_total = cons_qs.count()
        cons_completed = cons_qs.filter(completed_at__isnull=False).count()
        cons_cancelled = cons_qs.filter(cancelled_at__isnull=False).count()
        cons_planned = cons_total - cons_completed - cons_cancelled
        completion_rate = round((cons_completed / cons_total) * 100) if cons_total else 0

        today = timezone.now().date()
        recent_from = today - timedelta(days=30)
        previous_from = today - timedelta(days=60)
        recent_requests = req_qs.filter(created_at__date__gte=recent_from).count()
        previous_requests = req_qs.filter(
            created_at__date__gte=previous_from,
            created_at__date__lt=recent_from,
        ).count()
        request_trend_delta = recent_requests - previous_requests
        if request_trend_delta < 0:
            request_trend_text = 'Снижение числа обращений за последние 30 дней.'
        elif request_trend_delta > 0:
            request_trend_text = 'Рост числа обращений за последние 30 дней.'
        else:
            request_trend_text = 'Частота обращений стабильна.'

        positive_keywords = ('прогресс', 'улучш', 'стабил', 'справ', 'нормализ', 'спокойн')
        negative_keywords = ('тревог', 'стресс', 'конфликт', 'булл', 'агресс', 'депресс', 'паник', 'проблем')
        pos_q = Q()
        neg_q = Q()
        for word in positive_keywords:
            pos_q |= Q(result__icontains=word)
        for word in negative_keywords:
            neg_q |= Q(result__icontains=word)

        positive_result_signals = cons_qs.exclude(result__isnull=True).exclude(result__exact='').filter(pos_q).count()
        negative_result_signals = cons_qs.exclude(result__isnull=True).exclude(result__exact='').filter(neg_q).count()

        req_notes_qs = RequestNote.objects.filter(request__student_id=student.id)
        if user.role_name == 'psychologist':
            req_notes_qs = req_notes_qs.filter(
                Q(request__psychologist_id=user.id) | Q(request__psychologist_id__isnull=True)
            )
        if date_from:
            req_notes_qs = req_notes_qs.filter(created_at__date__gte=date_from)
        if date_to:
            req_notes_qs = req_notes_qs.filter(created_at__date__lte=date_to)

        pos_note_q = Q()
        neg_note_q = Q()
        for word in positive_keywords:
            pos_note_q |= Q(text__icontains=word)
        for word in negative_keywords:
            neg_note_q |= Q(text__icontains=word)
        positive_note_signals = req_notes_qs.filter(pos_note_q).count()
        negative_note_signals = req_notes_qs.filter(neg_note_q).count()

        # Отменённые консультации не считаем автоматически риском:
        # отмена может быть по нейтральным причинам (перенос, личные обстоятельства и т.п.).
        success_signals = cons_completed + positive_result_signals + positive_note_signals
        neutral_signals = cons_cancelled
        problem_signals = (
            req_status_map['new']
            + req_status_map['in_progress']
            + negative_result_signals
            + negative_note_signals
        )

        if cons_total == 0 and req_total <= 1:
            dynamics_level = 'Недостаточно данных'
            dynamics_comment = 'Пока недостаточно истории наблюдений для обоснованной оценки динамики.'
        elif completion_rate >= 70 and request_trend_delta <= 0 and success_signals >= problem_signals:
            dynamics_level = 'Положительная'
            dynamics_comment = 'Наблюдается положительная динамика: завершённых консультаций больше, чем риск-факторов.'
        elif request_trend_delta > 0 or problem_signals > success_signals:
            dynamics_level = 'Зона риска'
            dynamics_comment = 'Есть признаки сохранения/роста трудностей. Рекомендуется усилить сопровождение.'
        else:
            dynamics_level = 'Стабильная'
            if neutral_signals > 0:
                dynamics_comment = (
                    'Состояние без выраженного ухудшения. Есть отменённые консультации, '
                    'но они не трактуются как риск без дополнительных негативных факторов.'
                )
            else:
                dynamics_comment = 'Состояние без выраженного ухудшения, но требуется дальнейшее наблюдение.'

        months_ru = (
            '',
            'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
            'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь',
        )

        req_months = {
            (row['month'].year, row['month'].month): row['cnt']
            for row in req_qs.annotate(month=TruncMonth('created_at')).values('month').annotate(cnt=Count('id'))
            if row.get('month')
        }
        cons_months = {
            (row['month'].year, row['month'].month): {
                'completed': row.get('completed', 0) or 0,
                'cancelled': row.get('cancelled', 0) or 0,
            }
            for row in cons_qs.annotate(month=TruncMonth('date')).values('month').annotate(
                completed=Count('id', filter=Q(completed_at__isnull=False)),
                cancelled=Count('id', filter=Q(cancelled_at__isnull=False)),
            )
            if row.get('month')
        }

        month_keys = sorted(set(req_months.keys()) | set(cons_months.keys()))[-12:]
        chart_rows = []
        for year, month in month_keys:
            label = f'{months_ru[month]} {year}' if month < len(months_ru) else f'{month:02d}.{year}'
            chart_rows.append({
                'label': label,
                'requests': req_months.get((year, month), 0),
                'completed': cons_months.get((year, month), {}).get('completed', 0),
                'cancelled': cons_months.get((year, month), {}).get('cancelled', 0),
            })

        consultation_notes = Note.objects.filter(consultation__in=cons_qs).select_related('user').order_by('-created_at')[:8]
        request_notes = req_notes_qs.select_related('user').order_by('-created_at')[:8]
        recent_consultations = cons_qs.order_by('-date', '-start_time')[:10]

        query_params = self.request.GET.urlencode()
        back_url = reverse('consultations:report')
        if query_params:
            back_url = f'{back_url}?{query_params}'

        ctx.update({
            'student': student,
            'date_from': date_from,
            'date_to': date_to,
            'back_url': back_url,
            'req_total': req_total,
            'req_status_map': req_status_map,
            'cons_total': cons_total,
            'cons_completed': cons_completed,
            'cons_cancelled': cons_cancelled,
            'cons_planned': cons_planned,
            'completion_rate': completion_rate,
            'recent_requests': recent_requests,
            'previous_requests': previous_requests,
            'request_trend_delta': request_trend_delta,
            'request_trend_text': request_trend_text,
            'success_signals': success_signals,
            'neutral_signals': neutral_signals,
            'problem_signals': problem_signals,
            'dynamics_level': dynamics_level,
            'dynamics_comment': dynamics_comment,
            'chart_rows_json': json.dumps(chart_rows),
            'consultation_notes': consultation_notes,
            'request_notes': request_notes,
            'recent_consultations': recent_consultations,
        })
        return ctx


# ——— Экспорт отчётов ———

def _get_psychologist_querysets(user, date_from, date_to, status, student_id):
    'Queryset обращений и консультаций психолога с фильтрами. Обращения: созданные учащимися (без психолога) + назначенные этому психологу.'
    qs_req = Request.objects.filter(Q(psychologist_id=user.id) | Q(psychologist_id__isnull=True))
    qs_cons = Consultation.objects.select_related('request').filter(
        Q(request__psychologist_id=user.id)
        | Q(request_id__isnull=True)
        | Q(request__psychologist_id__isnull=True)
    )
    qs_req, qs_cons = _apply_report_filters(qs_req, qs_cons, date_from, date_to, status, student_id)
    return qs_req, qs_cons


def _get_students_report_data(qs_req, qs_cons):
    'Данные для отчёта «Обращения и консультации по учащимся».'
    student_ids = set(qs_req.values_list('student_id', flat=True))
    student_ids |= set(qs_cons.values_list('request__student_id', flat=True))
    student_ids |= set(qs_cons.values_list('students__id', flat=True))
    student_ids.discard(None)
    students_list = list(Student.objects.filter(pk__in=student_ids).order_by('last_name', 'first_name')) if student_ids else []
    result = []
    for s in students_list:
        req_cnt = qs_req.filter(student_id=s.id).count()
        cons_qs = qs_cons.filter(Q(request__student_id=s.id) | Q(students__id=s.id)).distinct()
        cons_cnt = cons_qs.count()
        last_cons = cons_qs.aggregate(m=Max('date'))['m']
        result.append({
            'student': s,
            'request_count': req_cnt,
            'consultation_count': cons_cnt,
            'last_consultation_date': last_cons,
        })
    return result


def _get_students_queryset_for_psychologist(user, date_from, date_to, status, student_id):
    'Учащиеся по фильтрам для отчёта психолога (свои обращения/консультации).'
    qs_req, qs_cons = _get_psychologist_querysets(user, date_from, date_to, status, student_id)
    student_ids = set(qs_req.values_list('student_id', flat=True))
    student_ids |= set(qs_cons.values_list('request__student_id', flat=True))
    student_ids |= set(qs_cons.values_list('students__id', flat=True))
    student_ids.discard(None)
    return Student.objects.filter(pk__in=student_ids).order_by('last_name', 'first_name') if student_ids else Student.objects.none()


def _get_consultations_queryset_for_admin(date_from, date_to, status, student_id):
    'Консультации по фильтрам для отчёта админа.'
    qs_req = Request.objects.all()
    qs_cons = Consultation.objects.select_related('request')
    qs_req, qs_cons = _apply_report_filters(qs_req, qs_cons, date_from, date_to, status, student_id)
    return qs_cons.select_related('request__student', 'request__psychologist', 'form').prefetch_related('students').order_by('-date')[:2000]


def _get_request_dynamics_data(qs_req):
    'Динамика обращений по месяцам: новые, в работе, завершённые, отменённые.'
    months_ru = ('', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь')
    month_counts = {}  # (year, month) -> {'req_new': N, 'req_in_progress': N, 'req_completed': N, 'req_cancelled': N}

    def _norm(m):
        if m is None:
            return None
        return (m.year, m.month) if hasattr(m, 'year') else None

    for d in qs_req.select_related('status').annotate(month=TruncMonth('created_at')).values('month', 'status__name').annotate(cnt=Count('id')):
        key = _norm(d['month'])
        if key:
            month_counts.setdefault(key, {'req_new': 0, 'req_in_progress': 0, 'req_completed': 0, 'req_cancelled': 0})
            status_name = d.get('status__name', '')
            if status_name == 'new':
                month_counts[key]['req_new'] = d['cnt']
            elif status_name == 'in_progress':
                month_counts[key]['req_in_progress'] = d['cnt']
            elif status_name == 'completed':
                month_counts[key]['req_completed'] = d['cnt']
            elif status_name == 'cancelled':
                month_counts[key]['req_cancelled'] = d['cnt']

    result = []
    for (year, month) in sorted(month_counts.keys()):
        r = month_counts[(year, month)]
        lbl = f"{months_ru[month]} {year}" if month < len(months_ru) else f"{month:02d}.{year}"
        req_total = r['req_new'] + r['req_in_progress'] + r['req_completed'] + r['req_cancelled']
        result.append({
            'label': lbl,
            'req_new': r['req_new'],
            'req_in_progress': r['req_in_progress'],
            'req_completed': r['req_completed'],
            'req_cancelled': r['req_cancelled'],
            'cnt': req_total
        })
    return result


def _get_consultation_dynamics_data(qs_cons):
    'Динамика консультаций по месяцам: завершённые, отменённые.'
    months_ru = ('', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь')
    month_counts = {}  # (year, month) -> {'cons_completed': N, 'cons_cancelled': N}

    def _norm(m):
        if m is None:
            return None
        return (m.year, m.month) if hasattr(m, 'year') else None

    # По месяцу даты консультации считаем завершённые и отменённые
    for d in qs_cons.annotate(month=TruncMonth('date')).values('month').annotate(
        cons_completed=Count('id', filter=Q(completed_at__isnull=False)),
        cons_cancelled=Count('id', filter=Q(cancelled_at__isnull=False))
    ):
        key = _norm(d['month'])
        if key:
            month_counts[key] = {
                'cons_completed': d['cons_completed'] or 0,
                'cons_cancelled': d['cons_cancelled'] or 0
            }

    result = []
    for (year, month) in sorted(month_counts.keys()):
        r = month_counts[(year, month)]
        lbl = f"{months_ru[month]} {year}" if month < len(months_ru) else f"{month:02d}.{year}"
        total = r['cons_completed'] + r['cons_cancelled']
        result.append({
            'label': lbl,
            'cons_completed': r['cons_completed'],
            'cons_cancelled': r['cons_cancelled'],
            'cnt': total
        })
    return result


def _students_chart_rows(data_list, limit=10):
    """Данные для диаграммы: топ учащихся по числу консультаций."""
    rows = []
    for row in sorted(data_list, key=lambda x: (x['consultation_count'], x['request_count']), reverse=True)[:limit]:
        rows.append({
            'label': row['student'].full_name,
            'requests': row['request_count'],
            'consultations': row['consultation_count'],
        })
    return rows


def _consultations_form_stats(consultations):
    """Агрегация консультаций по форме (индивидуальная/групповая)."""
    stats = {'Индивидуальная': 0, 'Групповая': 0, 'Другое': 0}
    for c in consultations:
        name = (getattr(getattr(c, 'form', None), 'name', '') or '').strip()
        if name == 'individual':
            stats['Индивидуальная'] += 1
        elif name == 'group':
            stats['Групповая'] += 1
        else:
            stats['Другое'] += 1
    return stats


def _autosize_worksheet_columns(ws, max_width=42):
    """Подбор ширины колонок для читаемого Excel."""
    from openpyxl.utils import get_column_letter

    for col_idx, column in enumerate(ws.columns, start=1):
        length = 0
        for cell in column:
            try:
                value = str(cell.value) if cell.value is not None else ''
            except Exception:
                value = ''
            if len(value) > length:
                length = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, length + 2), max_width)


def _style_excel_chart(chart, *, width=22, height=11, legend_pos='r', y_max=None):
    """Единый стиль диаграмм для читаемого отображения в Excel."""
    # Консервативный стиль лучше совместим с Excel/LibreOffice и не "переворачивает" оси.
    chart.style = 2
    chart.width = width
    chart.height = height
    if hasattr(chart, 'legend') and chart.legend is not None:
        chart.legend.position = legend_pos
    if hasattr(chart, 'y_axis') and chart.y_axis is not None:
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.orientation = 'minMax'
        chart.y_axis.majorUnit = 1
        if y_max is not None:
            chart.y_axis.scaling.max = max(1, y_max)
        chart.y_axis.crosses = 'min'
    if hasattr(chart, 'x_axis') and chart.x_axis is not None:
        chart.x_axis.tickLblPos = 'low'
        chart.x_axis.crosses = 'min'
    if hasattr(chart, 'gapWidth'):
        chart.gapWidth = 120
    if hasattr(chart, 'overlap'):
        chart.overlap = 0


def _configure_pdf_styles(styles, font_name):
    """Назначает шрифт с кириллицей всем базовым стилям PDF."""
    for style_name in ('Title', 'Heading1', 'Heading2', 'Heading3', 'Normal'):
        if style_name in styles:
            styles[style_name].fontName = font_name


def _pdf_period_label(date_from, date_to):
    if date_from or date_to:
        return f"Период: {date_from or '—'} — {date_to or '—'}"
    return 'Период: за всё время'


def _chart_max_value(series, extra=1):
    vals = [v for seq in series for v in seq]
    return max(vals or [0]) + extra


class ExportStudentsReportPDFView(PsychologistRequiredMixin, View):
    'Экспорт отчёта «Обращения и консультации по учащимся» в PDF.'
    def get(self, request):
        if request.user.role_name != 'psychologist':
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied('Доступно только психологу.')
        date_from, date_to, status, student_id = _report_filters(request)
        qs_req, qs_cons = _get_psychologist_querysets(request.user, date_from, date_to, status, student_id)
        data_list = _get_students_report_data(qs_req, qs_cons)

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.lib.units import cm
        from io import BytesIO

        pdf_font = _get_pdf_cyrillic_font()
        styles = getSampleStyleSheet()
        _configure_pdf_styles(styles, pdf_font)

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        elements = []
        elements.append(Paragraph('Обращения и консультации по учащимся', styles['Title']))
        elements.append(Paragraph(_pdf_period_label(date_from, date_to), styles['Normal']))
        tbl_data = [['№', 'Учащийся', 'Класс', 'Обращений', 'Консультаций', 'Дата последней консультации']]
        for i, row in enumerate(data_list, 1):
            last_d = row['last_consultation_date'].strftime('%d.%m.%Y') if row['last_consultation_date'] else '—'
            tbl_data.append([str(i), row['student'].full_name, row['student'].class_name or '—',
                             str(row['request_count']), str(row['consultation_count']), last_d])
        t = Table(tbl_data, colWidths=[25, 120, 50, 60, 80, 90])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), pdf_font),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E7E6E6')]),
        ]))
        elements.append(t)
        chart_rows = _students_chart_rows(data_list, limit=8)
        if chart_rows:
            elements.append(Paragraph('Топ учащихся по консультациям', styles['Normal']))
            drawing = Drawing(17 * cm, 9 * cm)
            chart = VerticalBarChart()
            chart.x, chart.y, chart.width, chart.height = 2 * cm, 1.8 * cm, 14 * cm, 5.5 * cm
            chart.data = [
                [r['requests'] for r in chart_rows],
                [r['consultations'] for r in chart_rows],
            ]
            chart.categoryAxis.categoryNames = [r['label'][:16] for r in chart_rows]
            chart.categoryAxis.labels.fontName = pdf_font
            chart.categoryAxis.labels.fontSize = 8
            chart.categoryAxis.labels.boxAnchor = 'n'
            chart.categoryAxis.labels.dy = -4
            chart.valueAxis.labels.fontName = pdf_font
            chart.valueAxis.labels.fontSize = 8
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = _chart_max_value(chart.data, extra=1)
            chart.bars[0].fillColor = colors.HexColor('#2563eb')
            chart.bars[1].fillColor = colors.HexColor('#16a34a')
            chart.barLabelFormat = '%d'
            drawing.add(chart)
            elements.append(drawing)
        doc.build(elements)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report_students.pdf"'
        return response


class ExportStudentsReportExcelView(PsychologistRequiredMixin, View):
    'Экспорт отчёта «Обращения и консультации по учащимся» в Excel.'
    def get(self, request):
        if request.user.role_name != 'psychologist':
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied('Доступно только психологу.')
        date_from, date_to, status, student_id = _report_filters(request)
        qs_req, qs_cons = _get_psychologist_querysets(request.user, date_from, date_to, status, student_id)
        data_list = _get_students_report_data(qs_req, qs_cons)

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.chart import BarChart, Reference
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = 'Обращения и консультации'
        ws.append(['№', 'Учащийся', 'Класс', 'Обращений', 'Консультаций', 'Дата последней консультации'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for i, row in enumerate(data_list, 1):
            last_d = row['last_consultation_date'].strftime('%d.%m.%Y') if row['last_consultation_date'] else '—'
            ws.append([i, row['student'].full_name, row['student'].class_name or '—',
                       row['request_count'], row['consultation_count'], last_d])
        _autosize_worksheet_columns(ws)

        chart_rows = _students_chart_rows(data_list, limit=10)
        if chart_rows:
            ws_chart = wb.create_sheet('График по учащимся')
            ws_chart.append(['Учащийся', 'Обращения', 'Консультации'])
            for r in chart_rows:
                ws_chart.append([r['label'], r['requests'], r['consultations']])
            for cell in ws_chart[1]:
                cell.font = Font(bold=True)
            _autosize_worksheet_columns(ws_chart)

            chart = BarChart()
            chart.title = 'Топ учащихся по консультациям'
            chart.y_axis.title = 'Количество'
            chart.x_axis.title = 'Учащиеся'
            data = Reference(ws_chart, min_col=2, max_col=3, min_row=1, max_row=len(chart_rows) + 1)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(chart_rows) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.grouping = 'clustered'
            _style_excel_chart(chart, width=20, height=11, legend_pos='b')
            ws_chart.add_chart(chart, 'E3')
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report_students.xlsx"'
        return response


class ExportDynamicsPDFView(PsychologistRequiredMixin, View):
    'Экспорт отчётов «Динамика обращений» и «Динамика консультаций» в PDF.'
    def get(self, request):
        if request.user.role_name != 'psychologist':
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied('Доступно только психологу.')
        date_from, date_to, status, student_id = _report_filters(request)
        qs_req, qs_cons = _get_psychologist_querysets(request.user, date_from, date_to, status, student_id)
        request_dynamics = _get_request_dynamics_data(qs_req)
        consultation_dynamics = _get_consultation_dynamics_data(qs_cons)

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.lib.units import cm
        from io import BytesIO

        pdf_font = _get_pdf_cyrillic_font()
        styles = getSampleStyleSheet()
        _configure_pdf_styles(styles, pdf_font)

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        elements = []
        elements.append(Paragraph('Динамика обращений и консультаций', styles['Title']))
        elements.append(Paragraph(_pdf_period_label(date_from, date_to), styles['Normal']))

        # 1) Динамика обращений
        elements.append(Paragraph('Динамика обращений', styles['Heading2'] if 'Heading2' in styles else styles['Normal']))
        tbl_req = [['Месяц', 'Новые', 'В работе', 'Завершённые', 'Отменённые', 'Всего']]
        for d in request_dynamics:
            tbl_req.append([d['label'], str(d['req_new']), str(d['req_in_progress']), str(d['req_completed']), str(d['req_cancelled']), str(d['cnt'])])
        t1 = Table(tbl_req)
        t1.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), pdf_font),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E7E6E6')]),
        ]))
        elements.append(t1)
        if request_dynamics:
            drawing1 = Drawing(17 * cm, 9 * cm)
            chart1 = VerticalBarChart()
            chart1.x, chart1.y, chart1.width, chart1.height = 2 * cm, 1.8 * cm, 14 * cm, 5.5 * cm
            chart1.data = [[d['req_new'] for d in request_dynamics], [d['req_in_progress'] for d in request_dynamics], [d['req_completed'] for d in request_dynamics], [d['req_cancelled'] for d in request_dynamics]]
            chart1.categoryAxis.categoryNames = [d['label'][:20] for d in request_dynamics]
            chart1.categoryAxis.labels.fontName = pdf_font
            chart1.categoryAxis.labels.fontSize = 8
            chart1.categoryAxis.labels.boxAnchor = 'n'
            chart1.categoryAxis.labels.dy = -4
            chart1.valueAxis.labels.fontName = pdf_font
            chart1.valueAxis.labels.fontSize = 8
            chart1.valueAxis.valueMin = 0
            chart1.valueAxis.valueMax = _chart_max_value(chart1.data, extra=1)
            chart1.bars[0].fillColor = colors.HexColor('#6c757d')
            chart1.bars[1].fillColor = colors.HexColor('#0d6efd')
            chart1.bars[2].fillColor = colors.HexColor('#198754')
            chart1.bars[3].fillColor = colors.HexColor('#ffc107')
            chart1.barLabelFormat = '%d'
            chart1.groupSpacing = 8
            chart1.barSpacing = 2
            drawing1.add(chart1)
            elements.append(Paragraph(' ', styles['Normal']))
            elements.append(drawing1)

        # 2) Динамика консультаций
        elements.append(Paragraph('Динамика консультаций', styles['Heading2'] if 'Heading2' in styles else styles['Normal']))
        tbl_cons = [['Месяц', 'Завершённые', 'Отменённые', 'Всего']]
        for d in consultation_dynamics:
            tbl_cons.append([d['label'], str(d['cons_completed']), str(d['cons_cancelled']), str(d['cnt'])])
        t2 = Table(tbl_cons)
        t2.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), pdf_font),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E7E6E6')]),
        ]))
        elements.append(t2)
        if consultation_dynamics:
            drawing2 = Drawing(17 * cm, 9 * cm)
            chart2 = VerticalBarChart()
            chart2.x, chart2.y, chart2.width, chart2.height = 2 * cm, 1.8 * cm, 14 * cm, 5.5 * cm
            chart2.data = [[d['cons_completed'] for d in consultation_dynamics], [d['cons_cancelled'] for d in consultation_dynamics]]
            chart2.categoryAxis.categoryNames = [d['label'][:20] for d in consultation_dynamics]
            chart2.categoryAxis.labels.fontName = pdf_font
            chart2.categoryAxis.labels.fontSize = 8
            chart2.categoryAxis.labels.boxAnchor = 'n'
            chart2.categoryAxis.labels.dy = -4
            chart2.valueAxis.labels.fontName = pdf_font
            chart2.valueAxis.labels.fontSize = 8
            chart2.valueAxis.valueMin = 0
            chart2.valueAxis.valueMax = _chart_max_value(chart2.data, extra=1)
            chart2.bars[0].fillColor = colors.HexColor('#198754')
            chart2.bars[1].fillColor = colors.HexColor('#6c757d')
            chart2.barLabelFormat = '%d'
            chart2.groupSpacing = 8
            chart2.barSpacing = 2
            drawing2.add(chart2)
            elements.append(Paragraph(' ', styles['Normal']))
            elements.append(drawing2)

        doc.build(elements)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report_dynamics.pdf"'
        return response


class ExportDynamicsExcelView(PsychologistRequiredMixin, View):
    'Экспорт отчётов «Динамика обращений» и «Динамика консультаций» в Excel.'
    def get(self, request):
        if request.user.role_name != 'psychologist':
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied('Доступно только психологу.')
        date_from, date_to, status, student_id = _report_filters(request)
        qs_req, qs_cons = _get_psychologist_querysets(request.user, date_from, date_to, status, student_id)
        request_dynamics = _get_request_dynamics_data(qs_req)
        consultation_dynamics = _get_consultation_dynamics_data(qs_cons)

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.chart import BarChart, Reference
        from io import BytesIO

        wb = Workbook()
        ws_req = wb.active
        ws_req.title = 'Динамика обращений'
        ws_req.append(['Месяц', 'Новые', 'В работе', 'Завершённые', 'Отменённые', 'Всего'])
        for cell in ws_req[1]:
            cell.font = Font(bold=True)
        for d in request_dynamics:
            ws_req.append([d['label'], d['req_new'], d['req_in_progress'], d['req_completed'], d['req_cancelled'], d['cnt']])
        _autosize_worksheet_columns(ws_req)

        if request_dynamics:
            bar_req = BarChart()
            bar_req.title = 'Динамика обращений'
            bar_req.y_axis.title = 'Количество'
            bar_req.x_axis.title = 'Месяц'
            data_req = Reference(ws_req, min_col=2, max_col=5, min_row=1, max_row=len(request_dynamics) + 1)
            cats_req = Reference(ws_req, min_col=1, min_row=2, max_row=len(request_dynamics) + 1)
            bar_req.add_data(data_req, titles_from_data=True)
            bar_req.set_categories(cats_req)
            bar_req.grouping = 'clustered'
            req_y_max = _chart_max_value([
                [d['req_new'] for d in request_dynamics],
                [d['req_in_progress'] for d in request_dynamics],
                [d['req_completed'] for d in request_dynamics],
                [d['req_cancelled'] for d in request_dynamics],
            ], extra=1)
            _style_excel_chart(bar_req, width=20, height=11, legend_pos='r', y_max=req_y_max)
            ws_req.add_chart(bar_req, 'H3')

        ws_cons = wb.create_sheet('Динамика консультаций')
        ws_cons.append(['Месяц', 'Завершённые', 'Отменённые', 'Всего'])
        for cell in ws_cons[1]:
            cell.font = Font(bold=True)
        for d in consultation_dynamics:
            ws_cons.append([d['label'], d['cons_completed'], d['cons_cancelled'], d['cnt']])
        _autosize_worksheet_columns(ws_cons)

        if consultation_dynamics:
            bar_cons = BarChart()
            bar_cons.title = 'Динамика консультаций'
            bar_cons.y_axis.title = 'Количество'
            bar_cons.x_axis.title = 'Месяц'
            data_cons = Reference(ws_cons, min_col=2, max_col=3, min_row=1, max_row=len(consultation_dynamics) + 1)
            cats_cons = Reference(ws_cons, min_col=1, min_row=2, max_row=len(consultation_dynamics) + 1)
            bar_cons.add_data(data_cons, titles_from_data=True)
            bar_cons.set_categories(cats_cons)
            bar_cons.grouping = 'clustered'
            cons_y_max = _chart_max_value([
                [d['cons_completed'] for d in consultation_dynamics],
                [d['cons_cancelled'] for d in consultation_dynamics],
            ], extra=1)
            _style_excel_chart(bar_cons, width=20, height=11, legend_pos='r', y_max=cons_y_max)
            ws_cons.add_chart(bar_cons, 'F3')

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report_dynamics.xlsx"'
        return response


class ExportWorkloadPDFView(PsychologistRequiredMixin, View):
    'Экспорт отчёта «Нагрузка школьного психолога» в PDF.'
    def get(self, request):
        if request.user.role_name != 'psychologist':
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied('Доступно только психологу.')
        date_from, date_to, status, student_id = _report_filters(request)
        qs_req, qs_cons = _get_psychologist_querysets(request.user, date_from, date_to, status, student_id)
        completed_cons = qs_cons.filter(completed_at__isnull=False)
        dur_agg = completed_cons.aggregate(total=Sum('duration'), avg=Avg('duration'))

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.lib.units import cm
        from io import BytesIO

        pdf_font = _get_pdf_cyrillic_font()
        styles = getSampleStyleSheet()
        _configure_pdf_styles(styles, pdf_font)

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        elements = []
        elements.append(Paragraph('Нагрузка школьного психолога', styles['Title']))
        elements.append(Paragraph(_pdf_period_label(date_from, date_to), styles['Normal']))
        tbl_data = [
            ['Показатель', 'Значение'],
            ['Количество обращений', str(qs_req.count())],
            ['Количество консультаций', str(qs_cons.count())],
            ['Суммарная длительность (мин)', str(dur_agg['total'] or 0)],
            ['Средняя длительность (мин)', str(round(dur_agg['avg'], 1) if dur_agg['avg'] is not None else 0)],
        ]
        t = Table(tbl_data)
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), pdf_font),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E7E6E6')]),
        ]))
        elements.append(t)
        form_stats = _consultations_form_stats(qs_cons)
        chart_items = [(k, v) for k, v in form_stats.items() if v > 0]
        if chart_items:
            elements.append(Paragraph('Распределение консультаций по форме', styles['Normal']))
            drawing = Drawing(17 * cm, 8 * cm)
            chart = VerticalBarChart()
            chart.x, chart.y, chart.width, chart.height = 2 * cm, 1.8 * cm, 14 * cm, 5 * cm
            chart.data = [[v for _, v in chart_items]]
            chart.categoryAxis.categoryNames = [k for k, _ in chart_items]
            chart.categoryAxis.labels.fontName = pdf_font
            chart.categoryAxis.labels.fontSize = 8
            chart.valueAxis.labels.fontName = pdf_font
            chart.valueAxis.labels.fontSize = 8
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = _chart_max_value(chart.data, extra=1)
            chart.bars[0].fillColor = colors.HexColor('#2563eb')
            chart.barLabelFormat = '%d'
            chart.groupSpacing = 8
            chart.barSpacing = 2
            drawing.add(chart)
            elements.append(drawing)
        doc.build(elements)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report_workload.pdf"'
        return response


# Старые имена для обратной совместимости (редирект на новый отчёт)
class ExportStudentsPDFView(PsychologistRequiredMixin, View):
    def get(self, request):
        return redirect('consultations:export_students_report_pdf' + ('?' + request.GET.urlencode() if request.GET else ''))


class ExportStudentsExcelView(PsychologistRequiredMixin, View):
    def get(self, request):
        return redirect('consultations:export_students_report_excel' + ('?' + request.GET.urlencode() if request.GET else ''))


class ExportConsultationsPDFView(AdminRequiredMixin, View):
    'Экспорт отчёта по консультациям в PDF (админ).'
    def get(self, request):
        date_from, date_to, status, student_id = _report_filters(request)
        consultations = _get_consultations_queryset_for_admin(date_from, date_to, status, student_id)

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.lib.units import cm
        from io import BytesIO

        pdf_font = _get_pdf_cyrillic_font()
        styles = getSampleStyleSheet()
        _configure_pdf_styles(styles, pdf_font)

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        elements = []
        elements.append(Paragraph('Отчёт по консультациям', styles['Title']))
        elements.append(Paragraph(_pdf_period_label(date_from, date_to), styles['Normal']))
        data = [['№', 'Дата', 'Время начала — окончания', 'Учащийся', 'Форма', 'Психолог', 'Результат']]
        for i, c in enumerate(consultations, 1):
            student_name = c.students_display()
            psych = c.request.psychologist if c.request_id else None
            psych_name = getattr(psych, 'username', '—') if psych else '—'
            result = (c.result or '')[:80] + ('…' if (c.result or '') and len(c.result or '') > 80 else '')
            time_str = c.time_display() or '—'
            data.append([str(i), str(c.date), time_str, student_name, c.form_display, psych_name, result])
        t = Table(data, colWidths=[25, 65, 70, 100, 80, 90, 100])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), pdf_font),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E7E6E6')]),
        ]))
        elements.append(t)
        form_stats = _consultations_form_stats(consultations)
        chart_items = [(k, v) for k, v in form_stats.items() if v > 0]
        if chart_items:
            elements.append(Paragraph('Распределение консультаций по форме', styles['Normal']))
            drawing = Drawing(17 * cm, 8 * cm)
            chart = VerticalBarChart()
            chart.x, chart.y, chart.width, chart.height = 2 * cm, 1.8 * cm, 14 * cm, 5 * cm
            chart.data = [[v for _, v in chart_items]]
            chart.categoryAxis.categoryNames = [k for k, _ in chart_items]
            chart.categoryAxis.labels.fontName = pdf_font
            chart.categoryAxis.labels.fontSize = 8
            chart.valueAxis.labels.fontName = pdf_font
            chart.valueAxis.labels.fontSize = 8
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = _chart_max_value(chart.data, extra=1)
            chart.bars[0].fillColor = colors.HexColor('#2563eb')
            chart.barLabelFormat = '%d'
            chart.groupSpacing = 8
            chart.barSpacing = 2
            drawing.add(chart)
            elements.append(drawing)
        doc.build(elements)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report_consultations.pdf"'
        return response


class ExportConsultationsExcelView(AdminRequiredMixin, View):
    'Экспорт отчёта по консультациям в Excel (админ).'
    def get(self, request):
        date_from, date_to, status, student_id = _report_filters(request)
        consultations = _get_consultations_queryset_for_admin(date_from, date_to, status, student_id)

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.chart import PieChart, Reference
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = 'Консультации'
        ws.append(['№', 'Дата', 'Время начала — окончания', 'Учащийся', 'Форма', 'Психолог', 'Результат'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for i, c in enumerate(consultations, 1):
            student_name = c.students_display()
            psych = c.request.psychologist if c.request_id else None
            psych_name = getattr(psych, 'username', '—') if psych else '—'
            time_str = c.time_display() or '—'
            ws.append([i, c.date, time_str, student_name, c.form_display, psych_name, (c.result or '')[:500]])
        _autosize_worksheet_columns(ws)

        form_stats = _consultations_form_stats(consultations)
        chart_items = [(k, v) for k, v in form_stats.items() if v > 0]
        if chart_items:
            ws_chart = wb.create_sheet('График по формам')
            ws_chart.append(['Форма', 'Количество'])
            for name, count in chart_items:
                ws_chart.append([name, count])
            for cell in ws_chart[1]:
                cell.font = Font(bold=True)
            _autosize_worksheet_columns(ws_chart)

            pie = PieChart()
            pie.title = 'Распределение консультаций по форме'
            data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(chart_items) + 1)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(chart_items) + 1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats)
            _style_excel_chart(pie, width=14, height=10, legend_pos='r')
            ws_chart.add_chart(pie, 'D3')
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report_consultations.xlsx"'
        return response


# ——— Личный кабинет учащегося ———

def _resolve_psychologist_for_student(student_id):
    """Подбирает психолога для чата учащегося."""
    recent_request = (
        Request.objects
        .select_related('psychologist')
        .filter(student_id=student_id, psychologist_id__isnull=False)
        .order_by('-created_at')
        .first()
    )
    if recent_request and recent_request.psychologist and recent_request.psychologist.role_name in ('psychologist', 'admin'):
        return recent_request.psychologist

    from users.models import User
    psychologist = User.objects.filter(role__name='psychologist', is_active=True).order_by('id').first()
    if psychologist:
        return psychologist
    return User.objects.filter(role__name='admin', is_active=True).order_by('id').first()


def _mark_chat_messages_read_for_user(chat_id, user_id):
    """Отмечает сообщения чата как прочитанные конкретным пользователем (персонально)."""
    unread_ids = list(
        ChatMessage.objects
        .filter(chat_id=chat_id)
        .exclude(author_id=user_id)
        .exclude(read_marks__user_id=user_id)
        .values_list('id', flat=True)
    )
    if not unread_ids:
        return
    now = timezone.now()
    ChatMessageRead.objects.bulk_create(
        [ChatMessageRead(message_id=mid, user_id=user_id, read_at=now) for mid in unread_ids],
        ignore_conflicts=True,
    )


class StudentChatView(StudentRequiredMixin, View):
    """Личный чат учащегося с психологом."""
    template_name = 'consultations/student_chat.html'

    def _get_chat(self, request):
        sid = getattr(request.user, 'student_id', None)
        if not sid:
            return None
        return (
            StudentPsychologistChat.objects
            .select_related('student', 'psychologist')
            .filter(student_id=sid)
            .first()
        )

    def _ensure_chat(self, request):
        chat = self._get_chat(request)
        if chat:
            return chat
        psychologist = _resolve_psychologist_for_student(request.user.student_id)
        if not psychologist:
            return None
        try:
            return StudentPsychologistChat.objects.create(
                student_id=request.user.student_id,
                psychologist_id=psychologist.pk,
            )
        except Exception:
            return self._get_chat(request)

    def _build_context(self, request, form=None):
        chat = self._ensure_chat(request)
        messages_qs = ChatMessage.objects.none()
        if chat:
            messages_qs = ChatMessage.objects.filter(chat_id=chat.pk).select_related('author').order_by('created_at')
            _mark_chat_messages_read_for_user(chat.pk, request.user.id)
        return {
            'chat': chat,
            'messages_list': messages_qs,
            'form': form or ChatMessageForm(),
            'has_profile': bool(getattr(request.user, 'student_id', None)),
        }

    def get(self, request):
        if not getattr(request.user, 'student_id', None):
            messages.error(request, 'Ваш аккаунт не привязан к карточке учащегося.')
            return redirect('consultations:student_dashboard')
        return render(request, self.template_name, self._build_context(request))

    def post(self, request):
        if not getattr(request.user, 'student_id', None):
            return redirect('consultations:student_dashboard')
        form = ChatMessageForm(request.POST)
        chat = self._ensure_chat(request)
        if not chat:
            messages.error(request, 'Сейчас нет доступного психолога для чата. Обратитесь к администратору.')
            return render(request, self.template_name, self._build_context(request, form=form))
        if not form.is_valid():
            return render(request, self.template_name, self._build_context(request, form=form))
        ChatMessage.objects.create(
            chat_id=chat.pk,
            author_id=request.user.pk,
            text=form.cleaned_data['text'],
        )
        chat.save(update_fields=['updated_at'])
        messages.success(request, 'Сообщение отправлено.')
        return redirect('consultations:student_chat')


class PsychologistChatListView(PsychologistRequiredMixin, ListView):
    """Список личных чатов учащихся для психолога."""
    model = StudentPsychologistChat
    template_name = 'consultations/psychologist_chat_list.html'
    context_object_name = 'chats'
    paginate_by = 30

    def get_queryset(self):
        qs = StudentPsychologistChat.objects.select_related('student', 'student__classroom', 'psychologist')
        if self.request.user.role_name == 'psychologist':
            qs = qs.filter(psychologist_id=self.request.user.id)
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(student__last_name__icontains=q)
                | Q(student__first_name__icontains=q)
            )
        unread_subquery = (
            ChatMessage.objects
            .filter(chat_id=OuterRef('pk'))
            .exclude(author_id=self.request.user.id)
            .exclude(read_marks__user_id=self.request.user.id)
            .values('chat_id')
            .annotate(cnt=Count('id', distinct=True))
            .values('cnt')[:1]
        )
        qs = qs.annotate(
            last_message_at=Max('messages__created_at'),
            unread_count=Coalesce(Subquery(unread_subquery, output_field=IntegerField()), Value(0)),
        ).order_by('-last_message_at', '-updated_at', '-created_at')
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search_q'] = self.request.GET.get('q', '')
        return ctx


class PsychologistChatDetailView(PsychologistRequiredMixin, View):
    """Детальный просмотр и отправка сообщений в чате учащегося."""
    template_name = 'consultations/psychologist_chat_detail.html'

    def _get_chat(self, request, pk):
        qs = StudentPsychologistChat.objects.select_related('student', 'student__classroom', 'psychologist')
        if request.user.role_name == 'psychologist':
            qs = qs.filter(psychologist_id=request.user.id)
        return get_object_or_404(qs, pk=pk)

    def _build_context(self, request, chat, form=None):
        messages_qs = ChatMessage.objects.filter(chat_id=chat.pk).select_related('author').order_by('created_at')
        _mark_chat_messages_read_for_user(chat.pk, request.user.id)
        return {
            'chat': chat,
            'messages_list': messages_qs,
            'form': form or ChatMessageForm(),
            'can_send': request.user.role_name == 'psychologist',
        }

    def get(self, request, pk):
        chat = self._get_chat(request, pk)
        return render(request, self.template_name, self._build_context(request, chat))

    def post(self, request, pk):
        if request.user.role_name != 'psychologist':
            messages.error(request, 'Администратор может только просматривать переписку.')
            return redirect('consultations:psychologist_chat_detail', pk=pk)
        chat = self._get_chat(request, pk)
        form = ChatMessageForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, self._build_context(request, chat, form=form))
        ChatMessage.objects.create(
            chat_id=chat.pk,
            author_id=request.user.pk,
            text=form.cleaned_data['text'],
        )
        chat.save(update_fields=['updated_at'])
        messages.success(request, 'Сообщение отправлено.')
        return redirect('consultations:psychologist_chat_detail', pk=chat.pk)


class StudentDashboardView(StudentRequiredMixin, TemplateView):
    'Главная страница для учащегося: приветствие и быстрые действия.'
    template_name = 'consultations/student_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        user.refresh_from_db()  # подтянуть актуальную привязку учащегося из БД
        student = getattr(user, 'student', None)
        ctx['student'] = student
        ctx['has_profile'] = student is not None
        if student:
            recent = Request.objects.filter(student_id=user.student_id).order_by('-created_at')[:5]
            ctx['recent_requests'] = recent
            ctx['request_count'] = Request.objects.filter(student_id=user.student_id).count()
            ctx['notifications'] = (
                StudentNotification.objects
                .filter(student_id=user.student_id)
                .select_related('consultation', 'request__status')
                .order_by('-created_at')[:30]
            )
        else:
            ctx['recent_requests'] = []
            ctx['request_count'] = 0
            ctx['notifications'] = []
        return ctx


class MyRequestListView(StudentRequiredMixin, ListView):
    'Список обращений учащегося (только свои).'
    model = Request
    template_name = 'consultations/my_request_list.html'
    context_object_name = 'requests'
    paginate_by = 10

    def get_queryset(self):
        if not getattr(self.request.user, 'student_id', None):
            return Request.objects.none()
        return (
            Request.objects
            .filter(student_id=self.request.user.student_id)
            .select_related('status')
            .order_by('-created_at')
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['has_profile'] = self.request.user.student_id is not None
        return ctx


class MyRequestCreateView(StudentRequiredMixin, View):
    """Учащийся подаёт обращение к психологу."""
    template_name = 'consultations/my_request_create.html'

    def get(self, request, *args, **kwargs):
        if not getattr(request.user, 'student_id', None):
            messages.error(request, 'Ваш аккаунт не привязан к карточке учащегося. Обратитесь к администратору.')
            return redirect('consultations:student_dashboard')
        return render(request, self.template_name, {'has_profile': True, 'form': MyRequestCreateForm()})

    def post(self, request, *args, **kwargs):
        if not getattr(request.user, 'student_id', None):
            return redirect('consultations:student_dashboard')
        form = MyRequestCreateForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'has_profile': True, 'form': form})
        status_new = RequestStatus.objects.filter(name='new').first()
        if not status_new:
            messages.error(request, 'В системе не настроен статус «Новое». Обратитесь к администратору.')
            return redirect('consultations:student_dashboard')
        request_obj = Request.objects.create(
            student_id=request.user.student_id,
            source=Request.SOURCE_STUDENT,
            status=status_new,
        )
        note_text = form.cleaned_data.get('note')
        if note_text:
            RequestNote.objects.create(
                request_id=request_obj.pk,
                user_id=request.user.pk,
                text=note_text,
            )
        messages.success(request, 'Обращение за консультацией отправлено. Психолог свяжется с вами.')
        return redirect('consultations:my_request_list')


class MyRequestDetailView(StudentRequiredMixin, DetailView):
    'Просмотр учащимся своего обращения (без редактирования/удаления).'
    model = Request
    template_name = 'consultations/my_request_detail.html'
    context_object_name = 'request_obj'

    def get_queryset(self):
        if not getattr(self.request.user, 'student_id', None):
            return Request.objects.none()
        return (
            Request.objects
            .filter(student_id=self.request.user.student_id)
            .select_related('status')
            .prefetch_related('consultations__form')
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['has_profile'] = self.request.user.student_id is not None
        req = ctx.get('request_obj')
        ctx['request_notes'] = (
            RequestNote.objects.filter(request_id=req.pk).select_related('user').order_by('created_at')
            if req else []
        )
        return ctx


class MyRequestCancelView(StudentRequiredMixin, View):
    'Отмена обращения учащимся (до статуса «Завершено»).'
    def post(self, request, pk):
        if not getattr(request.user, 'student_id', None):
            return redirect('consultations:student_dashboard')
        request_obj = get_object_or_404(
            Request.objects.filter(student_id=request.user.student_id),
            pk=pk
        )
        if request_obj.status.name == 'completed':
            messages.warning(request, 'Завершённое обращение отменить нельзя.')
            return redirect('consultations:my_request_detail', pk=pk)
        cancelled_status = RequestStatus.objects.filter(name='cancelled').first()
        if not cancelled_status:
            messages.error(request, 'В системе не найден статус «Отменено». Обратитесь к администратору.')
            return redirect('consultations:my_request_detail', pk=pk)
        request_obj.status = cancelled_status
        request_obj.save(update_fields=['status_id'])
        notify_request_status_changed(request_obj)
        messages.success(request, 'Обращение отменено.')
        return redirect('consultations:my_request_list')


class MyConsultationListView(StudentRequiredMixin, ListView):
    'Список консультаций учащегося (только свои).'
    model = Consultation
    template_name = 'consultations/my_consultation_list.html'
    context_object_name = 'consultations'
    paginate_by = 10

    def get_queryset(self):
        if not getattr(self.request.user, 'student_id', None):
            return Consultation.objects.none()
        return (
            Consultation.objects
            .filter(
                Q(request__student_id=self.request.user.student_id)
                | Q(students__id=self.request.user.student_id)
            )
            .select_related('form', 'request')
            .prefetch_related('consultation_students')
            .distinct()
            .order_by('-date', '-created_at')
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['has_profile'] = self.request.user.student_id is not None
        ctx['student_id'] = getattr(self.request.user, 'student_id', None)
        # Подтверждение участия текущего учащегося по каждой консультации
        sid = ctx['student_id']
        if sid and 'consultations' in ctx:
            for c in ctx['consultations']:
                cs = ConsultationStudent.objects.filter(consultation=c, student_id=sid).first()
                c.my_participation_confirmed_at = cs.participation_confirmed_at if cs else None
                c.my_participation_cancelled_at = cs.participation_cancelled_at if cs else None
        return ctx


class MyConsultationConfirmParticipationView(StudentRequiredMixin, View):
    'Подтверждение участия в консультации учащимся.'
    def post(self, request, pk):
        if not getattr(request.user, 'student_id', None):
            return redirect('consultations:student_dashboard')
        consultation = get_object_or_404(
            Consultation.objects.filter(
                Q(request__student_id=request.user.student_id)
                | Q(students__id=request.user.student_id)
            ).distinct(),
            pk=pk
        )
        if consultation.cancelled_at:
            messages.warning(request, 'Нельзя подтвердить участие в отменённой консультации.')
            return redirect('consultations:my_consultation_list')
        if consultation.completed_at:
            messages.warning(request, 'Нельзя подтвердить участие в уже завершённой консультации.')
            return redirect('consultations:my_consultation_list')
        cs = ConsultationStudent.objects.filter(
            consultation_id=pk,
            student_id=request.user.student_id
        ).first()
        if not cs:
            # Учащийся в консультации через request, но не в M2M — создаём запись через add
            consultation.students.add(request.user.student_id)
            cs = ConsultationStudent.objects.get(consultation_id=pk, student_id=request.user.student_id)
        if cs.participation_cancelled_at:
            messages.warning(request, 'Участие было окончательно отменено. Подтвердить снова нельзя.')
            return redirect('consultations:my_consultation_list')
        if cs.participation_confirmed_at:
            messages.info(request, 'Участие уже подтверждено.')
        else:
            cs.participation_confirmed_at = timezone.now()
            cs.save(update_fields=['participation_confirmed_at'])
            messages.success(request, 'Участие в консультации подтверждено.')
        return redirect('consultations:my_consultation_list')


class MyConsultationCancelParticipationView(StudentRequiredMixin, View):
    'Отмена участия в консультации учащимся: страница подтверждения и снятие подтверждения.'
    def get(self, request, pk):
        if not getattr(request.user, 'student_id', None):
            return redirect('consultations:student_dashboard')
        consultation = get_object_or_404(
            Consultation.objects.filter(
                Q(request__student_id=request.user.student_id)
                | Q(students__id=request.user.student_id)
            ).select_related('form').distinct(),
            pk=pk
        )
        if consultation.cancelled_at:
            messages.warning(request, 'Нельзя отменить участие в уже отменённой консультации.')
            return redirect('consultations:my_consultation_list')
        if consultation.completed_at:
            messages.warning(request, 'Нельзя отменить участие в уже завершённой консультации.')
            return redirect('consultations:my_consultation_list')
        cs = ConsultationStudent.objects.filter(
            consultation_id=pk,
            student_id=request.user.student_id
        ).first()
        if not cs:
            messages.info(request, 'Вы не записаны на эту консультацию.')
            return redirect('consultations:my_consultation_list')
        if cs.participation_cancelled_at:
            messages.info(request, 'Участие уже отменено.')
            return redirect('consultations:my_consultation_list')
        return render(
            request,
            'consultations/my_consultation_cancel_confirm.html',
            {'consultation': consultation}
        )

    def post(self, request, pk):
        if not getattr(request.user, 'student_id', None):
            return redirect('consultations:student_dashboard')
        consultation = get_object_or_404(
            Consultation.objects.filter(
                Q(request__student_id=request.user.student_id)
                | Q(students__id=request.user.student_id)
            ).distinct(),
            pk=pk
        )
        if consultation.cancelled_at:
            messages.warning(request, 'Нельзя отменить участие в уже отменённой консультации.')
            return redirect('consultations:my_consultation_list')
        if consultation.completed_at:
            messages.warning(request, 'Нельзя отменить участие в уже завершённой консультации.')
            return redirect('consultations:my_consultation_list')
        cs = ConsultationStudent.objects.filter(
            consultation_id=pk,
            student_id=request.user.student_id
        ).first()
        if not cs:
            messages.info(request, 'Вы не записаны на эту консультацию.')
            return redirect('consultations:my_consultation_list')
        if cs.participation_cancelled_at:
            messages.info(request, 'Участие уже отменено.')
            return redirect('consultations:my_consultation_list')
        cs.participation_confirmed_at = None
        cs.participation_cancelled_at = timezone.now()
        cs.save(update_fields=['participation_confirmed_at', 'participation_cancelled_at'])
        # Консультация получает статус «Отменена»
        consultation.cancelled_at = timezone.now()
        consultation.save(update_fields=['cancelled_at'])
        # Обращение по этой консультации тоже переводим в «Отменённые»
        if consultation.request_id:
            cancelled_status = RequestStatus.objects.filter(name='cancelled').first()
            if cancelled_status:
                consultation.request.status = cancelled_status
                consultation.request.save(update_fields=['status_id'])
                notify_request_status_changed(consultation.request)
        messages.success(request, 'Участие в консультации окончательно отменено. Консультация и обращение отменены.')
        return redirect('consultations:my_consultation_list')
